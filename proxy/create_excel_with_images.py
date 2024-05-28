import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import os
from PIL import Image as PILImage

# Загрузка данных из JSON файла
coins_data = pd.read_json('coins.json')
olympic_data = pd.read_json('images.json')

# Создание Excel файла
with pd.ExcelWriter('output.xlsx', engine='openpyxl') as writer:
    coins_data.to_excel(writer, index=False)

# Загрузка данных из Excel файла в DataFrame
df = pd.read_excel('output.xlsx')

# Загрузка рабочей книги Excel
wb = load_workbook('output.xlsx')
ws = wb.active

# Удаляем все текущие строки (если нужно очистить лист)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        ws.delete_rows(cell.row)

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 30
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 30
ws.column_dimensions['I'].width = 30
ws.column_dimensions['J'].width = 30
ws.cell(row=1, column=9, value='Olympic Games')
ws.cell(row=1, column=10, value='Olympic City')
# Установка жирного шрифта для текста в ячейке A1
ws['I1'].font = Font(bold=True)
# Центрирование текста в ячейке A1
ws['I1'].alignment = Alignment(horizontal='center', vertical='center')
ws['J1'].font = Font(bold=True)
ws['J1'].alignment = Alignment(horizontal='center', vertical='center')

# Переносим данные из DataFrame обратно в лист (опционально, если были изменения)
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
    # Задаем высоту для ячеек во второй строке
    ws.row_dimensions[r_idx].height = 80  # Высота в пунктах
    ws.cell(row=r_idx, column=1, value='')
    for c_idx, value in enumerate(row, 1):
        if c_idx in (1,2):
            value = ''
        ws.cell(row=r_idx, column=c_idx+1, value=value)

    olympic = olympic_data[olympic_data['code'] == value]  
 # Проверяем, что фильтрация вернула данные, и присваиваем значение из 'games'
    if not olympic.empty:
        games_value = olympic.iloc[0]['games']  # Берем значение игр из первой подходящей строки
        title_value = olympic.iloc[0]['title']
        src_value = 'c:/work/MY/sources/MyCoins/assets/' + olympic.iloc[0]['src']
        ws.cell(row=r_idx, column=c_idx + 2, value=games_value)  # Может потребоваться настроить номер столбца
        ws.cell(row=r_idx, column=c_idx + 3, value=title_value)
        # Загрузка изображения с помощью Pillow для определения исходных размеров
        with PILImage.open(src_value) as img:
            original_width, original_height = img.size

        # Вычисление новой ширины и высоты, сохраняющих пропорции (например, максимальная ширина или высота = 100 пикселей)
        max_size = 100
        ratio = min(max_size / original_width, max_size / original_height)
        new_width = int(original_width * ratio)
        new_height = int(original_height * ratio)

        # Загрузка изображения для вставки в Excel
        img = Image(src_value)
        img.width, img.height = new_width, new_height
        # Добавляем изображение в Excel файл (например, в столбец рядом с данными)
        ws.add_image(img, f'A{r_idx}')  

    # Путь к изображению из столбца 'src1'
    img_path1 = 'c:/work/MY/sources/MyCoins/assets/' + row[0]  # Предполагается, что 'src1' находится в первом столбце
    img_path2 = 'c:/work/MY/sources/MyCoins/assets/' + row[1]  # Предполагается, что 'src1' находится в первом столбце

    if os.path.exists(img_path1):
        img = Image(img_path1)
        img.width, img.height = 100, 100  # Установка размера изображения
        # Добавляем изображение в Excel файл (например, в столбец рядом с данными)
        ws.add_image(img, f'B{r_idx}')  
    if os.path.exists(img_path2):
        img = Image(img_path2)
        img.width, img.height = 100, 100  # Установка размера изображения
        # Добавляем изображение в Excel файл (например, в столбец рядом с данными)
        ws.add_image(img, f'C{r_idx}')  




# Сохраняем изменения в файл
wb.save('output_with_images.xlsx')



